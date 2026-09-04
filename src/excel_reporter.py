import pandas as pd
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers
)
from pathlib import Path
import logging
 
logger = logging.getLogger(__name__)

def create_excel_report(
    merged_df: pd.DataFrame,
    output_path: str = "reports/weather_report.xlsx"
) -> Path:
    """
    Create Excel with border styling.
    
    Features:
    - Different border styles (thin, thick, double)
    - Cell borders around tables
    - Alternating row colors
    """
    try:
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Styled Data"
        
        logger.info(f"Creating Excel with borders: {output_file}")
        
        # ===== DEFINE BORDER STYLES =====
        
        # Thick border for headers
        thick_border = Border(
            left=Side(style='thick', color='4472C4'),
            right=Side(style='thick', color='4472C4'),
            top=Side(style='thick', color='4472C4'),
            bottom=Side(style='thick', color='4472C4')
        )
        
        # Thin border for data
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        # ===== WRITE HEADERS =====
        
        headers = ['City', 'Latitude', 'Longitude', 'Date', 'Max Temp (°C)', 'Precipitation (mm)']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thick_border
        
        # ===== WRITE DATA WITH ALTERNATING COLORS =====
        
        light_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        for row_num, (_, row_data) in enumerate(merged_df.iterrows(), 2):
            # Alternate row colors
            row_fill = light_fill if row_num % 2 == 0 else white_fill
            
            # Create cells for each column
            cells_data = [
                row_data['city'],
                row_data['latitude'],
                row_data['longitude'],
                str(row_data['date']),
                row_data['max_temperature'],
                row_data['total_precipitation']
            ]
            
            for col_num, value in enumerate(cells_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Format numbers
                if col_num in [5, 6]:  # Temperature and precipitation columns
                    if col_num == 5:
                        cell.number_format = '0.0"°C"'
                    else:
                        cell.number_format = '0.0"mm"'
        
        # ===== SET COLUMN WIDTHS =====
        
        column_widths = {
            'A': 15, 'B': 12, 'C': 12,
            'D': 15, 'E': 15, 'F': 18
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # ===== FREEZE PANES =====
        
        ws.freeze_panes = "A2"
        
        # Save workbook
        wb.save(output_file)
        logger.info(f"Excel with borders created: {output_file}")
        print(f"✓ Excel with borders saved: {output_file}")
        
        return output_file
        
    except Exception as e:
        logger.error(f"Error creating Excel with borders: {e}")
        raise
 